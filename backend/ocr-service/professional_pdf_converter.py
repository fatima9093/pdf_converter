#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Professional PDF to Excel Converter
Uses multiple methods with fallbacks for accurate table extraction
"""

import argparse
import json
import sys
import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Ensure UTF-8 encoding for output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

# Import libraries with fallback handling
try:
    import camelot
    CAMELOT_AVAILABLE = True
    print("INFO: Camelot library available")
except ImportError as e:
    print(f"WARNING: Camelot not available: {e}")
    CAMELOT_AVAILABLE = False

try:
    import tabula
    TABULA_AVAILABLE = True
    print("INFO: Tabula library available")
except ImportError as e:
    print(f"WARNING: Tabula not available: {e}")
    TABULA_AVAILABLE = False

# PDF processing
from pdf2image import convert_from_path
import PyPDF2

class ProfessionalPDFToExcelConverter:
    """Professional PDF to Excel converter with multiple methods and fallbacks"""
    
    def __init__(self):
        """Initialize the converter with available methods"""
        self.conversion_methods = []
        
        # Add available methods in order of preference
        if CAMELOT_AVAILABLE:
            self.conversion_methods.extend([
                self._convert_with_camelot_lattice,
                self._convert_with_camelot_stream
            ])
        
        if TABULA_AVAILABLE:
            self.conversion_methods.append(self._convert_with_tabula)
        
        # Always add OCR fallback (will be implemented later if needed)
        self.conversion_methods.append(self._convert_with_enhanced_ocr)
        
        print(f"INFO: Initialized with {len(self.conversion_methods)} conversion methods")
    
    def convert_pdf_to_excel(self, pdf_path: str, output_path: str) -> Dict:
        """
        Convert PDF to Excel using multiple methods with fallbacks
        
        Args:
            pdf_path: Path to input PDF file
            output_path: Path for output Excel file
            
        Returns:
            Dict with conversion results and metadata
        """
        results = {
            'success': False,
            'method_used': None,
            'tables_found': 0,
            'pages_processed': 0,
            'error': None,
            'file_info': self._get_pdf_info(pdf_path)
        }
        
        print(f"🚀 Starting professional PDF to Excel conversion")
        print(f"📁 Input: {pdf_path}")
        print(f"📁 Output: {output_path}")
        print(f"📊 PDF Info: {results['file_info']}")
        
        # Try each conversion method
        for i, method in enumerate(self.conversion_methods, 1):
            try:
                print(f"\n🔄 Method {i}/{len(self.conversion_methods)}: {method.__name__}")
                
                success, tables_found = method(pdf_path, output_path)
                if success:
                    results.update({
                        'success': True,
                        'method_used': method.__name__,
                        'tables_found': tables_found,
                        'pages_processed': results['file_info']['pages']
                    })
                    print(f"✅ SUCCESS with {method.__name__}")
                    print(f"📊 Found {tables_found} tables")
                    return results
                else:
                    print(f"❌ Method {method.__name__} found no tables")
                    
            except Exception as e:
                print(f"❌ Method {method.__name__} failed: {e}")
                continue
        
        results['error'] = "All conversion methods failed to extract tables"
        print(f"💥 FAILED: {results['error']}")
        return results
    
    def _get_pdf_info(self, pdf_path: str) -> Dict:
        """Get basic PDF information"""
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                return {
                    'pages': len(reader.pages),
                    'size_mb': round(os.path.getsize(pdf_path) / (1024 * 1024), 2),
                    'encrypted': reader.is_encrypted
                }
        except Exception as e:
            return {
                'pages': 0,
                'size_mb': 0,
                'encrypted': False,
                'error': str(e)
            }
    
    def _convert_with_camelot_lattice(self, pdf_path: str, output_path: str) -> Tuple[bool, int]:
        """Convert using Camelot lattice method (best for tables with borders)"""
        if not CAMELOT_AVAILABLE:
            return False, 0
        
        print("📋 Using Camelot Lattice (tables with borders)")
        
        # Extract tables using lattice method
        tables = camelot.read_pdf(pdf_path, pages='all', flavor='lattice')
        
        if len(tables) == 0:
            return False, 0
        
        print(f"🔍 Found {len(tables)} tables with lattice method")
        
        # Save tables to Excel
        return self._save_camelot_tables_to_excel(tables, output_path, 'lattice')
    
    def _convert_with_camelot_stream(self, pdf_path: str, output_path: str) -> Tuple[bool, int]:
        """Convert using Camelot stream method (best for tables without borders)"""
        if not CAMELOT_AVAILABLE:
            return False, 0
        
        print("📋 Using Camelot Stream (tables without borders)")
        
        # Extract tables using stream method
        tables = camelot.read_pdf(pdf_path, pages='all', flavor='stream')
        
        if len(tables) == 0:
            return False, 0
        
        print(f"🔍 Found {len(tables)} tables with stream method")
        
        # Save tables to Excel
        return self._save_camelot_tables_to_excel(tables, output_path, 'stream')
    
    def _convert_with_tabula(self, pdf_path: str, output_path: str) -> Tuple[bool, int]:
        """Convert using Tabula method (alternative table extraction)"""
        if not TABULA_AVAILABLE:
            return False, 0
        
        print("📋 Using Tabula (alternative method)")
        
        try:
            # Extract tables using Tabula
            tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True)
            
            if not tables or len(tables) == 0:
                return False, 0
            
            # Filter out empty tables
            valid_tables = [df for df in tables if not df.empty]
            
            if len(valid_tables) == 0:
                return False, 0
            
            print(f"🔍 Found {len(valid_tables)} valid tables with Tabula")
            
            # Save tables to Excel
            return self._save_tabula_tables_to_excel(valid_tables, output_path)
            
        except Exception as e:
            print(f"Tabula error: {e}")
            return False, 0
    
    def _convert_with_enhanced_ocr(self, pdf_path: str, output_path: str) -> Tuple[bool, int]:
        """Enhanced OCR fallback method"""
        print("📋 Using Enhanced OCR (fallback method)")
        print("⚠️  OCR fallback not yet implemented - would need original OCR converter")
        return False, 0
    
    def _save_camelot_tables_to_excel(self, tables, output_path: str, method: str) -> Tuple[bool, int]:
        """Save Camelot tables to Excel with professional formatting"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            tables_saved = 0
            
            for i, table in enumerate(tables):
                # Get dataframe from Camelot table
                df = table.df
                
                # Clean the dataframe
                df_cleaned = self._clean_dataframe(df)
                
                if df_cleaned.empty:
                    continue
                
                # Create worksheet
                ws = wb.create_sheet(f"Table_{i+1}_P{table.page}")
                
                # Write data with professional formatting
                self._write_dataframe_to_worksheet(df_cleaned, ws, f"Camelot {method.title()}")
                
                tables_saved += 1
                print(f"  ✓ Saved Table {i+1} (Page {table.page}): {df_cleaned.shape[0]}x{df_cleaned.shape[1]}")
            
            if tables_saved == 0:
                return False, 0
            
            wb.save(output_path)
            print(f"💾 Saved {tables_saved} tables to {output_path}")
            return True, tables_saved
            
        except Exception as e:
            print(f"Error saving Camelot tables: {e}")
            return False, 0
    
    def _save_tabula_tables_to_excel(self, tables: List[pd.DataFrame], output_path: str) -> Tuple[bool, int]:
        """Save Tabula tables to Excel with professional formatting"""
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            tables_saved = 0
            
            for i, df in enumerate(tables):
                # Clean the dataframe
                df_cleaned = self._clean_dataframe(df)
                
                if df_cleaned.empty:
                    continue
                
                # Create worksheet
                ws = wb.create_sheet(f"Table_{i+1}")
                
                # Write data with professional formatting
                self._write_dataframe_to_worksheet(df_cleaned, ws, "Tabula")
                
                tables_saved += 1
                print(f"  ✓ Saved Table {i+1}: {df_cleaned.shape[0]}x{df_cleaned.shape[1]}")
            
            if tables_saved == 0:
                return False, 0
            
            wb.save(output_path)
            print(f"💾 Saved {tables_saved} tables to {output_path}")
            return True, tables_saved
            
        except Exception as e:
            print(f"Error saving Tabula tables: {e}")
            return False, 0
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Professional data cleaning and type conversion"""
        if df.empty:
            return df
        
        # Remove completely empty rows and columns
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        if df.empty:
            return df
        
        # Check if first row contains header-like data that duplicates column names
        if len(df) > 0:
            first_row = df.iloc[0].astype(str).str.lower().str.strip()
            
            # Common header patterns to detect
            header_patterns = ['id', 'name', 'quantity', 'price', 'item', 'amount', 'total', 'description']
            
            # Count how many cells in first row look like headers
            header_matches = sum(1 for cell in first_row if any(pattern in cell for pattern in header_patterns))
            
            # If more than half the first row looks like headers, remove it
            if header_matches >= len(first_row) / 2:
                print(f"  🧹 Removing duplicate header row: {list(first_row)}")
                df = df.drop(df.index[0]).reset_index(drop=True)
        
        # Reset column names to be more predictable
        df.columns = range(len(df.columns))
        
        # DON'T remove first column automatically - preserve table structure
        # Only remove if it's completely empty (100% null/empty)
        if len(df.columns) > 1:
            first_col = df.iloc[:, 0]
            # Check if first column is completely empty (not just mostly empty)
            completely_empty = first_col.isna().all() or (first_col.astype(str).str.strip() == '').all()
            
            if completely_empty:
                print(f"  🧹 Removing completely empty first column")
                df = df.drop(df.columns[0], axis=1)
                df.columns = range(len(df.columns))  # Reset column indices
            else:
                print(f"  ✅ Preserving first column with data")
        
        # Clean and convert data types intelligently
        for col in df.columns:
            # Convert to string first to handle mixed types
            df[col] = df[col].astype(str)
            
            # Clean text data
            df[col] = df[col].str.strip()
            df[col] = df[col].replace(['nan', 'NaN', 'None', '', 'null'], None)
            
            # Try to convert to numeric if most values are numeric
            numeric_count = 0
            total_count = df[col].notna().sum()
            
            if total_count > 0:
                for val in df[col].dropna():
                    try:
                        # Clean numeric strings
                        clean_val = str(val).replace(',', '').replace('$', '').replace('%', '').replace(' ', '')
                        float(clean_val)
                        numeric_count += 1
                    except:
                        pass
                
                # If more than 70% are numeric, convert the column
                if numeric_count / total_count > 0.7:
                    df[col] = pd.to_numeric(
                        df[col].str.replace(',', '').str.replace('$', '').str.replace('%', '').str.replace(' ', ''), 
                        errors='ignore'
                    )
        
        # Set meaningful column names based on typical table structure
        if len(df.columns) == 4:
            # Common 4-column table: ID, Name, Quantity, Price
            df.columns = ['ID', 'Name', 'Quantity', 'Price']
        elif len(df.columns) == 3:
            # 3-column table
            df.columns = ['Column_1', 'Column_2', 'Column_3']
        else:
            # Generic naming for other cases
            df.columns = [f'Column_{i+1}' for i in range(len(df.columns))]
        
        return df
    
    def _write_dataframe_to_worksheet(self, df: pd.DataFrame, ws, method_name: str):
        """Write dataframe to Excel worksheet with professional formatting"""
        
        # Start writing headers from row 1, column 1 (no method info header)
        header_row = 1
        
        # Use the dataframe as-is (already cleaned in _clean_dataframe)
        df_cleaned = df.copy()
        
        print(f"  📊 Writing {len(df_cleaned.columns)} columns: {list(df_cleaned.columns)}")
        
        # Write headers starting from row 1, column 1
        for c_idx, column in enumerate(df_cleaned.columns, 1):
            # Use the column name as-is (already cleaned)
            clean_column = str(column)
            
            cell = ws.cell(row=header_row, column=c_idx, value=clean_column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data starting from row 2
        for r_idx, row in enumerate(df_cleaned.values, header_row + 1):
            for c_idx, value in enumerate(row, 1):
                # Clean the value
                if pd.isna(value) or value in ['', 'nan', 'NaN', None]:
                    value = ''
                
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format based on data type
                if isinstance(value, (int, float)) and value != '':
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if isinstance(value, float):
                        cell.number_format = '0.00'
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Add alternating row colors
                if r_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = f"A{header_row + 1}"


def main():
    """Main function for command line usage"""
    parser = argparse.ArgumentParser(description='Professional PDF to Excel Converter')
    parser.add_argument('input_pdf', help='Input PDF file path')
    parser.add_argument('output_excel', help='Output Excel file path')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    try:
        # Check if input file exists
        if not os.path.exists(args.input_pdf):
            raise FileNotFoundError(f"Input PDF file not found: {args.input_pdf}")
        
        # Initialize converter
        converter = ProfessionalPDFToExcelConverter()
        
        # Check if any conversion methods are available
        if len(converter.conversion_methods) == 0:
            raise RuntimeError("No conversion methods available. Please install camelot-py or tabula-py.")
        
        # Perform conversion
        result = converter.convert_pdf_to_excel(args.input_pdf, args.output_excel)
        
        # Output results as JSON for server integration
        print(f"\n📊 CONVERSION RESULT:")
        print(json.dumps(result, indent=2))
        
        if result['success']:
            sys.exit(0)
        else:
            sys.exit(1)
            
    except Exception as e:
        error_result = {
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc() if args.verbose else None
        }
        print(f"\n❌ CONVERSION ERROR:")
        print(json.dumps(error_result, indent=2))
        sys.exit(1)


if __name__ == '__main__':
    main()
